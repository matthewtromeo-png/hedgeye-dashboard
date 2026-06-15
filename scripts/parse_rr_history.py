"""
parse_rr_history.py
───────────────────
Parses risk_range_tracker_excelworkbook.xlsx → builds RR_HISTORY JS object
→ patches the RR_HISTORY const in risk_range_dashboard.html in-place.

Run: python parse_rr_history.py
Called by: update_cowork.ps1 (before git push)
"""

import openpyxl, re, json, os, sys
from datetime import datetime

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT    = os.path.normpath(os.path.join(SCRIPT_DIR, '..'))
EXCEL_PATH   = r"C:\Users\matth\OneDrive\Desktop\Trading\risk_range_tracker_excelworkbook.xlsx"
DASHBOARD_HTML      = os.path.join(REPO_ROOT, 'project', 'risk_range_dashboard.html')
DASHBOARD_HTML_REPO = r'C:\repos\hedgeye-dashboard\project\risk_range_dashboard.html'
# CTX_PATH always writes to the git repo, not OneDrive
CTX_PATH         = r'C:\repos\hedgeye-dashboard\project\data\macro_context.json'
ONEDRIVE_CTX     = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'project', 'data', 'macro_context.json'))

def safe_read_ctx(write_path, fallback_path):
    """Read macro_context.json: try repo first, fall back to OneDrive if repo is missing/corrupt."""
    for label, p in [('repo', write_path), ('OneDrive', fallback_path)]:
        if not os.path.exists(p):
            continue
        try:
            with open(p, 'r', encoding='utf-8') as _f:
                return json.load(_f)
        except Exception as e:
            print(f"  [WARN] Could not read {label} JSON ({p}): {e}")
    return {}

# ── Parser ────────────────────────────────────────────────────────────────────
def parse_sheet_date(name):
    name = name.strip()
    m = re.match(r'([A-Za-z.]+)\s+(\d+)[.\s]+(\d{4})', name)
    if m:
        mon, day, yr = m.group(1).rstrip('.'), m.group(2), m.group(3)
        try:
            return datetime.strptime(f'{mon} {day} {yr}', '%b %d %Y').strftime('%Y-%m-%d')
        except Exception:
            pass
    return None

def parse_ticker_cell(val):
    if not val or not isinstance(val, str):
        return None, None
    m = re.match(r'^([A-Z0-9./^$]+)\s*\(?(BULLISH|BEARISH|NEUTRAL)?\)?', val.strip())
    if m and m.group(1) and len(m.group(1)) >= 2:
        return m.group(1), m.group(2)
    return None, None

def build_rr_history():
    if not os.path.exists(EXCEL_PATH):
        print(f"[WARN] Excel not found: {EXCEL_PATH}")
        return {}

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    skip = {'RiskRanges', 'Comparison'}
    history = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in skip:
            continue
        date_str = parse_sheet_date(sheet_name)
        if not date_str:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            ticker, signal = parse_ticker_cell(str(row[0]))
            if not ticker or len(ticker) < 2:
                continue
            if re.match(r'^[A-Z]$', ticker):
                continue
            lrr   = row[1] if len(row) > 1 else None
            trr   = row[2] if len(row) > 2 else None
            price = row[3] if len(row) > 3 else None
            try:
                lrr   = float(lrr)   if lrr   is not None else None
                trr   = float(trr)   if trr   is not None else None
                price = float(price) if price is not None else None
            except Exception:
                lrr = trr = price = None
            if lrr is None and trr is None:
                continue
            if ticker not in history:
                history[ticker] = {}
            if date_str not in history[ticker]:
                history[ticker][date_str] = {'l': lrr, 't': trr, 'p': price, 's': signal}

    result = {}
    for t, d in history.items():
        entries = [
            {'d': k, 'l': v['l'], 't': v['t'], 'p': v['p'], 's': v['s']}
            for k, v in sorted(d.items())
        ]
        if len(entries) >= 5:
            result[t] = entries
    return result

def patch_dashboard_html(rr_history):
    new_data = json.dumps(rr_history, separators=(',', ':'))
    new_const = f'const RR_HISTORY = {new_data};'
    placeholder = 'const RR_HISTORY = {};'

    targets = [
        ('OneDrive', DASHBOARD_HTML),
        ('Repo',     DASHBOARD_HTML_REPO),
    ]
    for label, fpath in targets:
        if not os.path.exists(fpath):
            print(f"[WARN] Dashboard HTML not found ({label}): {fpath}")
            continue
        with open(fpath, 'r', encoding='utf-8') as f:
            html = f.read()

        # ── Idempotent injection ──────────────────────────────────────────────
        # Step 1: Remove ALL existing RR_HISTORY declarations.
        #   Using count=1 + patched==html was a bug: when the JSON hasn't
        #   changed, the substitution produces the same string, the equality
        #   check is True, and the fallback injected a *second* declaration.
        #   That caused "SyntaxError: Identifier RR_HISTORY already declared."
        # Step 2: Inject exactly one new declaration at the inline <script> tag.
        cleaned, n_removed = re.subn(
            r'const RR_HISTORY\s*=\s*\{[\s\S]*?\};',
            '',
            html,
            flags=re.DOTALL,
        )
        if n_removed > 1:
            print(f"  [WARN] Removed {n_removed} duplicate RR_HISTORY declarations from {label}")

        # Always inject at the opening of the inline <script> block
        inline_open = '<script>'
        if inline_open not in cleaned:
            print(f"[WARN] Could not inject RR_HISTORY ({label}) — no inline <script> tag found")
            continue

        patched = cleaned.replace(inline_open, inline_open + '\n' + new_const, 1)
        print(f"Patched     : risk_range_dashboard.html ({label}) — {len(rr_history)} tickers  ({n_removed} old removed)")

        with open(fpath, 'w', encoding='utf-8') as f:
            f.write(patched)

        # Post-write validation: verify exactly one RR_HISTORY declaration
        count_after = len(re.findall(r'const RR_HISTORY\s*=', patched))
        if count_after != 1:
            print(f"  [ERROR] Post-write check failed ({label}): {count_after} RR_HISTORY declarations found (expected 1)")
            print(f"          This would cause a browser SyntaxError. Aborting.")
            raise RuntimeError(f"RR_HISTORY count={count_after} after injection")
        print(f"  Verified: {count_after} RR_HISTORY declaration in {label} HTML ✓")

def update_macro_context(rr_history):
    """Also store rr_history in macro_context.json for potential future use."""
    ctx = safe_read_ctx(CTX_PATH, ONEDRIVE_CTX)
    if not ctx:
        return

    ctx['rr_history'] = rr_history

    tmp = CTX_PATH + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(ctx, f, indent=2, ensure_ascii=False)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, CTX_PATH)
    print(f"Written     : macro_context.json  (rr_history — {len(rr_history)} tickers)")

def main():
    print("=== parse_rr_history.py ===")
    rr_history = build_rr_history()
    if not rr_history:
        print("[ERROR] No history data parsed — aborting")
        sys.exit(1)
    total = sum(len(v) for v in rr_history.values())
    print(f"Parsed      : {len(rr_history)} tickers, {total} entries")
    patch_dashboard_html(rr_history)
    update_macro_context(rr_history)
    print("Done.")

if __name__ == '__main__':
    main()
