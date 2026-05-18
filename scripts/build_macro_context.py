#!/usr/bin/env python3
"""
build_macro_context.py
======================
Stage 1: Excel / CSV sources (no API).
Stage 2: PDF extraction via Claude API (skipped with --stage1-only).

Stage 1 sources:
  - etf pro dash board\etf-pro-all-active-tickers-*.xlsx
  - HAM holdings\ETF_Holdings*.csv
  - RTA\real-time-alerts-history-*.csv
  - risk_range_tracker_excelworkbook.xlsx

Stage 2 PDF sources:
  DAILY   (cache 2 days,  newest 1 file each)
    macro_show, msr, sss, momo, early_look, portfolio,
    btc, call_summary, macro_show_notes
  WEEKLY  (cache 14 days, newest 2 files each)
    investing_ideas, founders_choice, from_desk
  SPECIAL (no expiry,     all files from last 35 days, one combined call)
    macro_research

Output: project/data/macro_context.json
"""

import argparse
import base64
import csv
import hashlib
import json
import os
import re
import sys
import tempfile
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.  Run: pip install openpyxl")
    sys.exit(1)

# ── PATHS ─────────────────────────────────────────────────────────────────────
HEDGEYE_BASE = Path(r"C:\Users\matth\OneDrive\Desktop\Trading\hedgeye")
REPO_ROOT    = Path(__file__).resolve().parent.parent
RR_WORKBOOK  = Path(r"C:\Users\matth\OneDrive\Desktop\Trading\risk_range_tracker_excelworkbook.xlsx")
OUTPUT_PATH  = REPO_ROOT / "project" / "data" / "macro_context.json"

# ── HELPERS ───────────────────────────────────────────────────────────────────

def newest_file(folder: Path, glob_pattern: str) -> Path | None:
    """Return most recently modified file in folder matching glob_pattern, or None."""
    matches = sorted(folder.glob(glob_pattern), key=lambda p: p.stat().st_mtime)
    if not matches:
        print(f"  [glob] No matches for {folder / glob_pattern}")
        return None
    print(f"  [glob] Found {len(matches)} match(es) in {folder}")
    print(f"  [glob] Using: {matches[-1].name}")
    return matches[-1]


def warn(msg: str) -> None:
    print(f"  [WARN] {msg}", file=sys.stderr)


def cell_float(v) -> float | None:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return None if f != f else f
    if isinstance(v, str):
        s = v.strip().replace(',', '').rstrip('%')
        try:
            return float(s) if s else None
        except ValueError:
            return None
    return None


def iso_date(v) -> str:
    """Normalise various date representations to YYYY-MM-DD string."""
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, str):
        s = v.strip()[:10]
        return s
    return ''


# ── ETF PRO ───────────────────────────────────────────────────────────────────

def read_etf_pro() -> dict:
    print("\n── ETF Pro ──")
    path = newest_file(HEDGEYE_BASE / "etf pro dash board", "etf-pro-all-active-tickers-*.xlsx")
    if not path:
        warn("etf pro: no file found")
        return {"etf_rerank": [], "active_longs": [], "active_shorts": [], "_source": None}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    print(f"  [debug] Total raw rows (including header): {len(rows)}")
    if not rows:
        warn("etf pro: empty workbook")
        return {"etf_rerank": [], "active_longs": [], "active_shorts": [], "_source": path.name}

    # Find the header row dynamically — it contains 'Ticker' somewhere in first 5 rows
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        if any(str(c or '').strip().lower() == 'ticker' for c in row):
            header_idx = i
            break
    print(f"  [debug] Header found at row index {header_idx}: {[str(c or '') for c in rows[header_idx]]}")
    header = [str(c).strip() if c else '' for c in rows[header_idx]]

    def col(name):
        try:
            return header.index(name)
        except ValueError:
            return None

    i_etf   = col('ETF')
    i_class = col('Asset Class')
    i_call  = col('Call')
    i_tick  = col('Ticker')
    i_date  = col('Date Added')
    i_price = col('Last Price')
    i_days  = col('Days Held')

    rerank, longs, shorts = [], [], []
    seen_tickers = set()
    today = date.today()

    for row in rows[header_idx + 1:]:
        def get(i, r=row):
            return r[i] if i is not None and i < len(r) else None

        ticker = str(get(i_tick) or '').strip()
        call   = str(get(i_call) or '').strip().upper()
        # Skip blank rows, header repeats, and non-Long/Short rows
        if not ticker or ticker.lower() == 'ticker' or call not in ('LONG', 'SHORT'):
            continue
        if ticker in seen_tickers:
            print(f"  [debug] Duplicate ticker skipped: {ticker}")
            continue
        seen_tickers.add(ticker)

        rerank.append(ticker)

        added_raw = get(i_date)
        days_held = get(i_days)
        if days_held is None and added_raw:
            try:
                added = datetime.strptime(str(added_raw)[:10], '%Y-%m-%d').date()
                days_held = (today - added).days
            except (ValueError, TypeError):
                days_held = None

        entry = {
            "ticker":      ticker,
            "etf":         str(get(i_etf) or '').strip(),
            "asset_class": str(get(i_class) or '').strip(),
            "date_added":  iso_date(added_raw),
            "last_price":  cell_float(get(i_price)),
            "days_held":   int(days_held) if days_held is not None else None,
        }
        if call == 'LONG':
            longs.append(entry)
        else:
            shorts.append(entry)

    print(f"  [debug] After filter: {len(rerank)} unique tickers ({len(longs)} long, {len(shorts)} short)")
    return {
        "etf_rerank":    rerank,
        "active_longs":  longs,
        "active_shorts": shorts,
        "_source":       path.name,
    }


# ── HAM HOLDINGS ─────────────────────────────────────────────────────────────

def read_ham_holdings() -> dict:
    print("\n── HAM Holdings ──")
    path = newest_file(HEDGEYE_BASE / "HAM holdings", "ETF_Holdings*.csv")
    if not path:
        warn("HAM holdings: no file found")
        return {"ham_holdings": [], "_source": None}

    by_ticker: dict = {}

    with open(path, newline='', encoding='utf-8-sig') as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            ticker = (row.get('StockTicker') or '').strip()
            if not ticker or ticker == 'Cash&Other':
                continue
            if (row.get('MoneyMarketFlag') or '').strip().upper() == 'Y':
                continue
            if '-TRS-' in ticker:
                continue

            raw_weight = (row.get('Weightings') or '0').replace('%', '').strip()
            try:
                weight = float(raw_weight) / 100
            except ValueError:
                continue
            if not weight or weight <= 0:
                continue

            account = (row.get('Account') or '').strip()
            name    = (row.get('SecurityName') or '').strip()

            raw_mv = (row.get('MarketValue') or '').replace(',', '').strip()
            try:
                market_value = float(raw_mv)
            except ValueError:
                market_value = None

            if ticker not in by_ticker:
                by_ticker[ticker] = {
                    "ticker":       ticker,
                    "name":         name,
                    "total_weight": 0.0,
                    "accounts":     {},
                    "market_value": 0.0,
                }
            entry = by_ticker[ticker]
            entry["total_weight"] = round(entry["total_weight"] + weight, 6)
            entry["accounts"][account] = round(
                entry["accounts"].get(account, 0.0) + weight, 6
            )
            if market_value:
                entry["market_value"] = round(
                    (entry["market_value"] or 0.0) + market_value, 2
                )
            if not entry["name"] and name:
                entry["name"] = name

    holdings = sorted(by_ticker.values(), key=lambda x: x["total_weight"], reverse=True)
    return {"ham_holdings": holdings, "_source": path.name}


# ── RTA TRADE HISTORY ─────────────────────────────────────────────────────────

def _parse_duration(raw: str) -> str:
    s = raw.lower()
    if 'tail'  in s: return 'TAIL'
    if 'trend' in s: return 'TREND'
    return 'TRADE'


def _parse_date(raw: str) -> date | None:
    """Parse RTA date strings in any of these formats:
      '2026-05-11 15:01:17 -0400'  (new CSV format)
      '2026-05-11'                 (ISO date only)
      '5/11/2026 15:01'            (old US datetime format)
      '5/11/2026'                  (old US date only)
    """
    s = raw.strip()
    # ISO format — detected by YYYY-MM-DD shape; slice [:10] strips time/tz
    if len(s) >= 10 and s[4:5] == '-' and s[7:8] == '-':
        try:
            return datetime.strptime(s[:10], '%Y-%m-%d').date()
        except ValueError:
            pass
    # US M/D/YYYY formats (legacy)
    for fmt in ('%m/%d/%Y %H:%M', '%m/%d/%Y'):
        try:
            return datetime.strptime(s[:len(fmt)], fmt).date()
        except ValueError:
            continue
    return None


def read_rta_trades() -> dict:
    print("\n── RTA ──")
    path = newest_file(HEDGEYE_BASE / "RTA", "real-time-alerts-history-*.csv")
    if not path:
        warn("RTA: no file found")
        return {"rta": {"recent_trades": [], "stats": {}, "recently_traded_tickers": []}, "_source": None}

    today      = date.today()
    cutoff_90d = today - timedelta(days=90)
    cutoff_60d = today - timedelta(days=60)
    print(f"  [debug] today={today}  cutoff_90d={cutoff_90d}  cutoff_60d={cutoff_60d}")

    all_closed: list = []
    skipped_open = skipped_date_err = total_rows = 0

    with open(path, newline='', encoding='utf-8-sig') as fh:
        reader = csv.DictReader(fh)
        first = True
        for row in reader:
            if first:
                print(f"  [debug] CSV columns: {list(row.keys())}")
                print(f"  [debug] First row sample: Close Date={row.get('Close Date')!r}  Close Price={row.get('Close Price')!r}")
                first = False
            total_rows += 1

            close_date_raw  = (row.get('Close Date')  or '').strip()
            close_price_raw = (row.get('Close Price') or '').strip()
            if not close_date_raw or not close_price_raw:
                skipped_open += 1
                continue

            close_date = _parse_date(close_date_raw)
            if close_date is None:
                skipped_date_err += 1
                continue

            close_date_str = close_date.strftime('%Y-%m-%d')
            symbol   = (row.get('Symbol')   or '').strip()
            position = (row.get('Position') or '').strip().lower()
            if not symbol:
                continue

            open_date_obj = _parse_date(row.get('Open Date') or '')
            open_date_str = open_date_obj.strftime('%Y-%m-%d') if open_date_obj else ''

            try:
                open_price = float((row.get('Open Price') or '').strip())
            except ValueError:
                open_price = None

            try:
                close_price = float(close_price_raw)
            except ValueError:
                close_price = None

            try:
                realized_return = float((row.get('Realized Return') or '').strip())
            except ValueError:
                realized_return = None

            duration = _parse_duration(row.get('Duration') or '')

            all_closed.append({
                "symbol":           symbol,
                "position":         position,
                "open_date":        open_date_str,
                "open_price":       open_price,
                "close_date":       close_date_str,
                "close_price":      close_price,
                "realized_return":  realized_return,
                "duration":         duration,
                "_close_date_obj":  close_date,
            })

    recent = [t for t in all_closed if t["_close_date_obj"] >= cutoff_90d]
    recent.sort(key=lambda x: x["close_date"], reverse=True)

    recently_traded = sorted({
        t["symbol"] for t in all_closed if t["_close_date_obj"] >= cutoff_60d
    })

    for t in recent:
        del t["_close_date_obj"]

    longs  = [t for t in recent if t["position"] == "long"  and t["realized_return"] is not None]
    shorts = [t for t in recent if t["position"] == "short" and t["realized_return"] is not None]

    def win_rate(trades):
        if not trades: return None
        return round(sum(1 for t in trades if t["realized_return"] > 0) / len(trades), 4)

    def avg_return(trades):
        if not trades: return None
        return round(sum(t["realized_return"] for t in trades) / len(trades), 4)

    stats = {
        "win_rate_long":    win_rate(longs),
        "win_rate_short":   win_rate(shorts),
        "avg_return_long":  avg_return(longs),
        "avg_return_short": avg_return(shorts),
        "total_trades_90d": len(recent),
    }

    return {
        "rta": {
            "recent_trades":           recent,
            "stats":                   stats,
            "recently_traded_tickers": recently_traded,
        },
        "_source": path.name,
    }


# ── RISK RANGE LEVELS ─────────────────────────────────────────────────────────

def read_risk_range_levels() -> dict:
    if not RR_WORKBOOK.exists():
        warn(f"risk range workbook not found: {RR_WORKBOOK}")
        return {"levels": {}, "_source": None}

    scripts_dir = Path(__file__).resolve().parent
    if str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))

    try:
        import import_official_levels as iol
    except ImportError as e:
        warn(f"Could not import import_official_levels: {e}")
        return {"levels": {}, "_source": RR_WORKBOOK.name}

    wb = openpyxl.load_workbook(RR_WORKBOOK, read_only=True, data_only=True)
    all_rows = []
    year_hint = date.today().year

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name in iol.SKIP_SHEETS:
            continue
        try:
            if ws.sheet_state in ('hidden', 'veryHidden'):
                continue
        except AttributeError:
            pass
        sheet_date = iol.sheet_name_to_date(sheet_name, year_hint)
        if not sheet_date:
            continue
        rows, _ = iol.parse_sheet(ws, sheet_date)
        all_rows.extend(rows)

    wb.close()

    latest_by_ticker: dict = {}
    for row in all_rows:
        t = row['ticker']
        if t not in latest_by_ticker or row['date'] > latest_by_ticker[t]['date']:
            latest_by_ticker[t] = row

    levels = {
        t: {
            "close":  r['close'],
            "lrr":    r['buy'],
            "trr":    r['sell'],
            "signal": r['signal'],
        }
        for t, r in latest_by_ticker.items()
        if r['buy'] or r['sell'] or r['close']
    }

    return {"levels": levels, "_source": RR_WORKBOOK.name}


# ── STAGE 2: PDF EXTRACTION VIA CLAUDE API ────────────────────────────────────
#
# PDF_SOURCES  — regular sources (daily or weekly, 1–2 files each)
# MACRO_RESEARCH_CFG — special: all files from last 35 days in one combined call
#
# Cache key for regular sources  = pipe-joined sorted filenames of the N files used
# Cache key for macro_research   = 12-char MD5 hash of sorted filenames
# Cache is valid when:           key matches AND age of last run < cache_days
#                                (macro_research has no cache_days → never expires on key match)

PDF_SOURCES = [
    # ── DAILY — cache 2 days, newest 1 file ───────────────────────────────────
    {
        "key":        "macro_show",
        "folder":     HEDGEYE_BASE / "macro show slides",
        "glob":       "HE_TMS_*.pdf",
        "cache_days": 2,
        "read_n":     1,
        "label":      "macro_show (daily)",
        "prompt": (
            "Extract these fields from the Hedgeye Macro Show presentation.\n"
            "Return ONLY valid JSON (null for any field not found):\n"
            "{\n"
            '  "quad": {"monthly": <int 1-4>, "quarterly": <int 1-4>, "depth": <"SHALLOW"|"MODERATE"|"DEEP">},\n'
            '  "vix": {"current": <float>, "lrr": <float>, "trr": <float>, "bucket": <"INVESTABLE"|"CAUTION"|"CRASH">},\n'
            '  "cpi_nowcast": <float>,\n'
            '  "growth_roc": <"ACCELERATING"|"DECELERATING">,\n'
            '  "inflation_roc": <"ACCELERATING"|"DECELERATING">,\n'
            '  "quad_sequence": <string like "2-2-2-3" showing next 4 quarters>,\n'
            '  "high_beta_1m": <float, 1-month return % for high-beta factor bucket>,\n'
            '  "low_beta_1m": <float, 1-month return % for low-beta factor bucket>,\n'
            '  "keith_commentary": [<up to 5 key bullet points as strings>]\n'
            "}"
        ),
    },
    {
        "key":        "msr",
        "folder":     HEDGEYE_BASE / "Market situation report",
        "glob":       "*.pdf",
        "cache_days": 2,
        "read_n":     1,
        "label":      "msr (daily)",
        "prompt": (
            "Extract these fields from the Hedgeye Market Situation Report (MSR) or Weekly Game Plan.\n"
            "Return ONLY valid JSON (null for any field not found):\n"
            "{\n"
            '  "gamma_exposure": <"POSITIVE"|"NEGATIVE"|"NEUTRAL">,\n'
            '  "systematic_flow": <"BUYING"|"SELLING"|"NEUTRAL">,\n'
            '  "pv_band": <"OVERBOUGHT"|"OVERSOLD"|"NEUTRAL">,\n'
            '  "strategic_allocation": <"RISK_ON"|"RISK_OFF"|"NEUTRAL">,\n'
            '  "spx_upper_pv": <float, upper price/volume level for SPX>,\n'
            '  "spx_lower_pv": <float, lower price/volume level for SPX>,\n'
            '  "gex_flip": <float, GEX zero or flip level for SPX>,\n'
            '  "gvt": <float, gamma velocity threshold>,\n'
            '  "realized_vol_10d": <float, 10-day realized volatility %>,\n'
            '  "spx_last": <float, last SPX price>,\n'
            '  "resistance": <float, key near-term resistance level>,\n'
            '  "support": <float, key near-term support level>\n'
            "}"
        ),
    },
    {
        "key":        "sss",
        "folder":     HEDGEYE_BASE / "signal strength list",
        "glob":       "*.pdf",
        "cache_days": 2,
        "read_n":     1,
        "label":      "sss (daily)",
        "prompt": (
            "Extract these fields from the Hedgeye Signal Strength Score (SSS) PDF.\n"
            "Return ONLY valid JSON (null for any field not found):\n"
            "{\n"
            '  "count": <int, total tickers on the list>,\n'
            '  "added": [<ticker symbols added this week>],\n'
            '  "removed": [<ticker symbols removed this week>],\n'
            '  "tickers": [<every ticker on the list in order of appearance>]\n'
            "}\n\n"
            "STRICT TICKER RULES — only include genuine US equity ticker symbols:\n"
            "  • 1–5 uppercase letters only, optionally followed by a single digit (e.g. NVDA, SPY, BRK, QQQ)\n"
            "  • DO NOT include: lowercase text, partial words, column headers, page numbers,\n"
            "    or strings containing $, %, ., /, -, spaces, or any other special character\n"
            "  • DO NOT include strings longer than 6 characters\n"
            "  • When in doubt, omit the entry\n"
        ),
    },
    {
        "key":        "momo",
        "folder":     HEDGEYE_BASE / "Momentum Stock Tracker",
        "glob":       "*.pdf",
        "cache_days": 2,
        "read_n":     1,
        "label":      "momo (daily)",
        "prompt": (
            "Extract risk range and signal data for every ticker in the Hedgeye Momentum Stock Tracker.\n"
            "Return ONLY valid JSON as a flat object keyed by ticker (null for missing values):\n"
            "{\n"
            '  "<TICKER>": {"lrr": <float>, "trr": <float>, "signal": <"BULLISH"|"BEARISH"|"NEUTRAL">, "close": <float>}\n'
            "}"
        ),
    },
    {
        "key":        "early_look",
        "folder":     HEDGEYE_BASE / "Early look pdfs",
        "glob":       "*.pdf",
        "cache_days": 2,
        "read_n":     1,
        "label":      "early_look (daily)",
        "prompt": (
            "Extract these fields from the Hedgeye Early Look PDF.\n"
            "Return ONLY valid JSON (null for any field not found):\n"
            "{\n"
            '  "key_levels": {\n'
            '    "<TICKER>": {"lrr": <float>, "trr": <float>, "signal": <"BULLISH"|"BEARISH"|"NEUTRAL">}\n'
            "  },\n"
            '  "signal_changes": [\n'
            '    {"ticker": <string>, "from": <"BULLISH"|"BEARISH"|"NEUTRAL">, "to": <"BULLISH"|"BEARISH"|"NEUTRAL">}\n'
            "  ],\n"
            '  "vix_level": <float>,\n'
            '  "keith_notes": [<up to 5 key observations or trade notes from Keith>]\n'
            "}"
        ),
    },
    {
        "key":        "portfolio",
        "folder":     HEDGEYE_BASE / "Portfolio solutions",
        "glob":       "*.pdf",
        "cache_days": 2,
        "read_n":     1,
        "label":      "portfolio (daily)",
        "prompt": (
            "Extract portfolio positions and movers from this Hedgeye Portfolio Solutions PDF.\n"
            "Return ONLY valid JSON (null for any field not found):\n"
            "{\n"
            '  "positions": [\n'
            '    {"ticker": <string>, "weight": <float 0-1>, "daily_move": <string e.g. "+1.2%">, '
            '"signal": <"BULLISH"|"BEARISH"|"NEUTRAL">}\n'
            "  ],\n"
            '  "top_movers": [\n'
            '    {"ticker": <string>, "move": <string e.g. "+8%">}\n'
            "  ],\n"
            '  "rerank_date": <string YYYY-MM-DD>\n'
            "}"
        ),
    },
    {
        "key":        "btc",
        "folder":     HEDGEYE_BASE / "BTC trend tracker",
        "glob":       "*.pdf",
        "cache_days": 2,
        "read_n":     1,
        "label":      "btc (daily)",
        "prompt": (
            "Extract risk range and signal data for all crypto assets from this Hedgeye BTC Trend Tracker PDF.\n"
            "Return ONLY valid JSON as a flat object keyed by symbol (null for missing values):\n"
            "{\n"
            '  "<SYMBOL>": {"lrr": <float>, "trr": <float>, "signal": <"BULLISH"|"BEARISH"|"NEUTRAL">}\n'
            "}"
        ),
    },
    {
        "key":        "call_summary",
        "folder":     HEDGEYE_BASE / "The call summaries",
        "glob":       "*.pdf",
        "cache_days": 2,
        "read_n":     1,
        "label":      "call_summary (daily)",
        "prompt": (
            "Extract key data from this Hedgeye macro call summary.\n"
            "Return ONLY valid JSON (null for any field not found):\n"
            "{\n"
            '  "date": <string YYYY-MM-DD>,\n'
            '  "quad": <int 1-4>,\n'
            '  "key_points": [<up to 5 most important takeaways from the call>],\n'
            '  "trades_mentioned": [<any specific trade ideas or position changes discussed>],\n'
            '  "keith_outlook": <string, 1-2 sentence summary of Keith\'s overall market view>,\n'
            '  "ticker_mentions": {\n'
            '    "<TICKER>": {\n'
            '      "pod1_revenue": <"ACCELERATING"|"DECELERATING"|null>,\n'
            '      "eps_trend": <"BEAT"|"MISS"|"IN_LINE"|null>,\n'
            '      "gross_margin": <"EXPANDING"|"CONTRACTING"|null>,\n'
            '      "key_thesis": <one sentence string|null>,\n'
            '      "catalysts": [<string>],\n'
            '      "signal": <"BULLISH"|"BEARISH"|"NEUTRAL"|null>\n'
            "    }\n"
            "  }\n"
            "}"
        ),
    },
    {
        "key":        "macro_show_notes",
        "folder":     HEDGEYE_BASE / "macro show slides" / "macro show reports",
        "glob":       "*.pdf",
        "cache_days": 2,
        "read_n":     1,
        "label":      "macro_show_notes (daily)",
        "prompt": (
            "Extract key takeaways from this Hedgeye Macro Show summary notes PDF.\n"
            "Return ONLY valid JSON (null for missing fields):\n"
            "{\n"
            '  "key_points": [<up to 5 most important takeaways>],\n'
            '  "positioning_changes": [<position adds, removes, or trims mentioned>],\n'
            '  "keith_watching": [<catalysts, price levels, or events Keith is watching>]\n'
            "}"
        ),
    },

    # ── WEEKLY — cache 14 days, newest 2 files ─────────────────────────────────
    {
        "key":        "investing_ideas",
        "folder":     HEDGEYE_BASE / "Investing Ideas",
        "glob":       "*.pdf",
        "cache_days": 14,
        "read_n":     2,
        "label":      "investing_ideas (weekly)",
        "prompt": (
            "Extract all current long and short positions from this Hedgeye Investing Ideas newsletter.\n"
            "If two issues are provided, use the more recent one for current positions.\n"
            "Return ONLY valid JSON (null for missing fields):\n"
            "{\n"
            '  "longs": {\n'
            '    "<TICKER>": {"lrr": <float lower risk range>, "trr": <float top risk range>, "thesis": <1-2 sentence summary>}\n'
            "  },\n"
            '  "shorts": {\n'
            '    "<TICKER>": {"lrr": <float>, "trr": <float>, "thesis": <1-2 sentence summary>}\n'
            "  }\n"
            "}"
        ),
    },
    {
        "key":         "founders_choice",
        "folder":      HEDGEYE_BASE / "Founders Choice",
        "glob":        "*.pdf",
        "cache_days":  14,
        "read_n":      2,
        "name_filter": ["Founder", "Position Monitor"],
        "label":       "founders_choice (weekly)",
        "prompt": (
            "Extract all sector long and short stock picks from this Hedgeye Founders Choice PDF.\n"
            "Return ONLY valid JSON (null if not found):\n"
            "{\n"
            '  "<sector e.g. industrials>": {"longs": [<ticker strings>], "shorts": [<ticker strings>]},\n'
            '  "ticker_mentions": {\n'
            '    "<TICKER>": {\n'
            '      "pod1_revenue": <"ACCELERATING"|"DECELERATING"|null>,\n'
            '      "eps_trend": <"BEAT"|"MISS"|"IN_LINE"|null>,\n'
            '      "gross_margin": <"EXPANDING"|"CONTRACTING"|null>,\n'
            '      "key_thesis": <one sentence string|null>,\n'
            '      "catalysts": [<string>],\n'
            '      "signal": <"BULLISH"|"BEARISH"|"NEUTRAL"|null>\n'
            "    }\n"
            "  }\n"
            "}"
        ),
    },
    {
        "key":        "from_desk",
        "folder":     HEDGEYE_BASE / "From the desk",
        "glob":       "*.pdf",
        "cache_days": 14,
        "read_n":     2,
        "label":      "from_desk (weekly)",
        "prompt": (
            "Extract key market commentary from this Hedgeye 'From the Desk' newsletter.\n"
            "Return ONLY valid JSON (null for any field not found):\n"
            "{\n"
            '  "macro_view": <string, 1-2 sentence summary of the macro thesis>,\n'
            '  "key_points": [<up to 5 key takeaways>],\n'
            '  "tickers_mentioned": [<all ticker symbols specifically mentioned>],\n'
            '  "risk_tone": <"RISK_ON"|"RISK_OFF"|"NEUTRAL">\n'
            "}"
        ),
    },
]

# Special — all PDFs from last 35 days sent as one combined API call
MACRO_RESEARCH_CFG = {
    "key":         "macro_research",
    "folder":      HEDGEYE_BASE / "macro research",
    "glob":        "*.pdf",
    "window_days": 35,
    "label":       "macro_research (combined, 35-day window)",
    "prompt": (
        "The following documents are recent Hedgeye macro research PDFs. They may include:\n"
        "  - Monthly Inflation Nowcast (CPI forecasts and upside/downside bands)\n"
        "  - Quarterly Macro Themes deck (quad sequence, GDP/CPI forecasts by quarter)\n"
        "  - Mid-Quarter Update (revised quad calls)\n"
        "  - Levels reference (SPX and other key ticker risk ranges)\n\n"
        "Extract ALL of the following fields across the provided documents.\n"
        "Return ONLY valid JSON (null for any field not found):\n"
        "{\n"
        '  "cpi_nowcast":    <float, current CPI nowcast %>,\n'
        '  "cpi_upside":     <float, upside CPI scenario %>,\n'
        '  "cpi_downside":   <float, downside CPI scenario %>,\n'
        '  "cpi_trend":      <"ACCELERATING"|"DECELERATING">,\n'
        '  "quarterly_quad": <int 1-4, current quarterly quad>,\n'
        '  "monthly_quad":   <int 1-4, current monthly quad>,\n'
        '  "quad_sequence":  <string like "2-2-2-3" — look for "Quad Count X-X-X-X" in the Mid-Quarter Update or Quarterly Themes deck; these are sequential monthly quads forward in time>,\n'
        '  "forward_quads": {\n'
        '    "<QNYY e.g. 2Q26>": {"quad": <int>, "gdp_qoq": <float %>, "cpi_yoy": <float %>}\n'
        "  },\n"
        '  "key_themes": [<up to 5 highest-conviction macro themes as strings>],\n'
        '  "levels_reference": {\n'
        '    "<TICKER>": {"lrr": <float>, "trr": <float>, "signal": <"BULLISH"|"BEARISH"|"NEUTRAL">}\n'
        "  }\n"
        "}"
    ),
}


# ── PDF CACHE HELPERS ─────────────────────────────────────────────────────────

def _newest_n(folder: Path, glob_pattern: str, n: int,
              name_filter: list[str] | None = None) -> list[Path]:
    """Return the n most recently modified files matching the pattern (silent).

    If name_filter is provided, only files whose name contains at least one of
    the filter strings (case-insensitive) are considered.
    """
    matches = sorted(folder.glob(glob_pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if name_filter:
        matches = [p for p in matches
                   if any(f.lower() in p.name.lower() for f in name_filter)]
    return matches[:n]


def _source_id(paths: list[Path]) -> str:
    """Cache key: 'name@mtime' entries — detects filename changes AND in-place updates."""
    return "|".join(sorted(f"{p.name}@{int(p.stat().st_mtime)}" for p in paths))


def _files_hash(paths: list[Path]) -> str:
    """12-char MD5 of 'name@mtime' pairs — compact key for multi-file sources."""
    combined = "|".join(sorted(f"{p.name}@{int(p.stat().st_mtime)}" for p in paths))
    return hashlib.md5(combined.encode()).hexdigest()[:12]


def _cache_valid(key: str, source_id: str, existing: dict | None) -> bool:
    """True if the cached source_id matches (same filename AND same mtime).
    Time-based expiry removed — re-extraction is driven by file changes only."""
    if not existing:
        return False
    return existing.get("sources_used", {}).get(key) == source_id


# ── PDF SPLITTING HELPERS ─────────────────────────────────────────────────────

def _pdf_page_count(path: Path) -> int:
    """Return number of pages in a PDF, or 0 if pypdf is unavailable/file unreadable."""
    try:
        import pypdf
        return len(pypdf.PdfReader(str(path)).pages)
    except Exception:
        return 0


def _split_pdf_pages(pdf_path: Path, chunk_size: int = 50) -> list[Path]:
    """Split a PDF into chunk_size-page temp files; caller must delete them.

    Returns a list of temp Paths (one per chunk, in page order).
    Falls back to [pdf_path] unchanged if pypdf is missing.
    """
    try:
        import pypdf
    except ImportError:
        warn("pypdf not installed — cannot split large PDF; run: pip install pypdf")
        return [pdf_path]

    reader = pypdf.PdfReader(str(pdf_path))
    total  = len(reader.pages)
    chunks: list[Path] = []

    for start in range(0, total, chunk_size):
        end    = min(start + chunk_size, total)
        writer = pypdf.PdfWriter()
        for i in range(start, end):
            writer.add_page(reader.pages[i])
        tmp = Path(tempfile.mktemp(suffix=f'_{pdf_path.stem}_p{start+1}-{end}.pdf'))
        with open(tmp, 'wb') as fh:
            writer.write(fh)
        print(f"  [split] pages {start+1}–{end} → {tmp.name}")
        chunks.append(tmp)

    return chunks


# ── CLAUDE API CALL ───────────────────────────────────────────────────────────

def _is_rate_limit(exc: Exception) -> bool:
    return getattr(exc, 'status_code', None) == 429 or 'RateLimitError' in type(exc).__name__


def _call_claude_pdf(
    client,
    paths: list[Path],
    prompt: str,
    label: str,
    sleep_after: int = 3,
    fallback: dict | None = None,
) -> dict | None:
    """Send one or more PDFs as document blocks, return parsed JSON or fallback on 429.

    Retry policy: on 429, sleep 30s and retry once. If still rate-limited,
    return fallback (the existing cached value) instead of None so we don't
    overwrite valid data with null.
    """
    total_kb = sum(p.stat().st_size for p in paths) // 1024
    print(f"  [API] {label}: {len(paths)} file(s), {total_kb} KB")
    for p in paths:
        print(f"        {p.name}")

    content = []
    for path in paths:
        pdf_b64 = base64.standard_b64encode(path.read_bytes()).decode('utf-8')
        content.append({
            "type": "document",
            "source": {
                "type":       "base64",
                "media_type": "application/pdf",
                "data":       pdf_b64,
            },
        })
    content.append({"type": "text", "text": prompt})

    def _do_call() -> dict | None:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2048,
            system=(
                "You are a financial data extractor. Extract ONLY the requested fields from "
                "this Hedgeye research document. Return ONLY valid JSON, no other text."
            ),
            messages=[{"role": "user", "content": content}],
        )
        raw = msg.content[0].text.strip()
        # Extract the JSON object — strip markdown fences and any surrounding prose
        try:
            raw = raw[raw.index('{') : raw.rindex('}') + 1]
        except ValueError:
            pass
        # Fix trailing commas before ] or } (common model output artifact)
        raw = re.sub(r',\s*([}\]])', r'\1', raw)
        result = json.loads(raw)
        print(f"  [API] OK — {len(result)} top-level keys")
        return result

    try:
        result = _do_call()
        return result
    except json.JSONDecodeError as e:
        warn(f"{label}: JSON parse failed — {e}")
        return None
    except Exception as e:
        if _is_rate_limit(e):
            warn(f"{label}: 429 rate limit — sleeping 30s and retrying once")
            time.sleep(30)
            try:
                result = _do_call()
                return result
            except json.JSONDecodeError as e2:
                warn(f"{label}: JSON parse failed on retry — {e2}")
                return None
            except Exception as e2:
                if _is_rate_limit(e2):
                    warn(f"{label}: 429 on retry — using cached fallback")
                    return fallback
                warn(f"{label}: API error on retry — {e2}")
                return fallback
        warn(f"{label}: API error — {e}")
        return None
    finally:
        time.sleep(sleep_after)  # rate-limit guard between API calls


# ── MERGE HELPER FOR MACRO RESEARCH ──────────────────────────────────────────

def _merge_macro_research(results: list[dict]) -> dict:
    """Merge per-file macro_research extractions.

    Scalars (cpi_nowcast, etc.): last non-null value wins (oldest→newest order).
    Dicts (forward_quads, levels_reference): union; newer file wins on key conflicts.
    Lists (key_themes): concatenate unique values, cap at 5.
    """
    merged: dict = {}
    scalar_keys = ('cpi_nowcast', 'cpi_upside', 'cpi_downside', 'cpi_trend',
                   'quarterly_quad', 'monthly_quad', 'quad_sequence')
    for r in results:
        for k in scalar_keys:
            if r.get(k) is not None:
                merged[k] = r[k]
        for dict_key in ('forward_quads', 'levels_reference'):
            if r.get(dict_key):
                merged.setdefault(dict_key, {}).update(r[dict_key])
        if r.get('key_themes'):
            existing_themes = merged.get('key_themes', [])
            for theme in r['key_themes']:
                if theme not in existing_themes:
                    existing_themes.append(theme)
            merged['key_themes'] = existing_themes[:5]
    return merged


# ── EXTRACT ALL PDFs ──────────────────────────────────────────────────────────

def extract_pdf_data(existing: dict | None, force_pdf: bool) -> dict:
    """Stage 2: extract structured data from PDFs via Claude API, with per-source caching."""
    print("\n── PDF Extraction (Stage 2) ──")

    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        warn("ANTHROPIC_API_KEY not set — skipping PDF extraction")
        return {"pdf": {}, "_pdf_sources": {}, "_cache_stats": {}}

    try:
        import anthropic
    except ImportError:
        warn("anthropic not installed — run: pip install anthropic")
        return {"pdf": {}, "_pdf_sources": {}, "_cache_stats": {}}

    client       = anthropic.Anthropic(api_key=api_key)
    results      = {}
    sources_used = {}
    cache_stats  = {}   # key → "cached" | "fresh (N files)" | "missing"

    MAX_PDF_MB = 10.0  # ~13MB base64; well under the API body limit (~20MB)

    # ── Regular sources (daily + weekly) ──────────────────────────────────────
    for src in PDF_SOURCES:
        key        = src["key"]
        cache_days = src["cache_days"]

        paths = _newest_n(src["folder"], src["glob"], src["read_n"], src.get("name_filter"))
        if not paths:
            warn(f"{src['label']}: no PDF found")
            results[key]      = None
            sources_used[key] = None
            cache_stats[key]  = "missing"
            continue

        # Size guard — skip rather than sending an oversized payload
        total_mb = sum(p.stat().st_size for p in paths) / 1_048_576
        if total_mb > MAX_PDF_MB:
            warn(f"{src['label']}: file too large ({total_mb:.1f}MB), skipping")
            results[key]      = None
            sources_used[key] = _source_id(paths)
            cache_stats[key]  = f"too large ({total_mb:.1f}MB)"
            continue

        source_id         = _source_id(paths)
        sources_used[key] = source_id

        cached_val = (existing or {}).get("pdf", {}).get(key)
        if not force_pdf and cached_val is not None and _cache_valid(key, source_id, existing):
            results[key]     = cached_val
            cache_stats[key] = "cached"
            print(f"  {src['label']:<36} cached  ({source_id[:60]})")
            continue

        results[key]     = _call_claude_pdf(client, paths, src["prompt"], src["label"], fallback=cached_val)
        cache_stats[key] = f"fresh ({len(paths)} file{'s' if len(paths) > 1 else ''})"
        if key == "macro_show":
            print("  [rate-limit] Sleeping 60s after macro_show...")
            time.sleep(60)

    # ── Post-process SSS: strip PDF-artifact ticker strings ──────────────────
    # Applies whether result came from cache or fresh extraction.
    _valid_ticker = re.compile(r'^[A-Z]{1,5}[0-9]?$')
    sss_result = results.get('sss')
    if sss_result and isinstance(sss_result.get('tickers'), list):
        raw = sss_result['tickers']
        clean = [t for t in raw if _valid_ticker.match(t)]
        bad   = [t for t in raw if not _valid_ticker.match(t)]
        if bad:
            print(f"  [sss] Removed {len(bad)} invalid ticker(s): {bad}")
        for fld in ('added', 'removed'):
            if isinstance(sss_result.get(fld), list):
                sss_result[fld] = [t for t in sss_result[fld] if _valid_ticker.match(t)]
        sss_result['tickers'] = clean
        sss_result['count']   = len(clean)
        results['sss'] = sss_result

    # ── Macro research: each file individually, merge results ─────────────────
    # Window reduced to 14 days; cap at 4 files; one API call per file.
    # Results are merged oldest-to-newest so newer files win on conflicts.
    cfg    = MACRO_RESEARCH_CFG
    cutoff = datetime.now() - timedelta(days=14)
    recent_pdfs = sorted(
        [p for p in cfg["folder"].glob(cfg["glob"])
         if datetime.fromtimestamp(p.stat().st_mtime) >= cutoff],
        key=lambda p: p.stat().st_mtime,
    )[:4]  # oldest first so newest overwrites on merge; hard cap at 4

    if not recent_pdfs:
        warn(f"{cfg['label']}: no PDFs found in last 14 days")
        results["macro_research"]      = None
        sources_used["macro_research"] = None
        cache_stats["macro_research"]  = "missing"
    else:
        source_id                      = _files_hash(recent_pdfs)
        sources_used["macro_research"] = source_id

        cached_val = (existing or {}).get("pdf", {}).get("macro_research")
        if not force_pdf and cached_val is not None and _cache_valid("macro_research", source_id, existing):
            results["macro_research"]     = cached_val
            cache_stats["macro_research"] = "cached"
            print(f"  {cfg['label']:<36} cached  ({len(recent_pdfs)} files, hash {source_id})")
        else:
            per_file: list[dict] = []
            for pdf_path in recent_pdfs:
                page_count = _pdf_page_count(pdf_path)
                if page_count > 95:
                    print(f"  [split] {pdf_path.name}: {page_count} pages — splitting into 50-page chunks")
                    chunk_paths = _split_pdf_pages(pdf_path, chunk_size=50)
                    chunk_results: list[dict] = []
                    try:
                        for chunk_path in chunk_paths:
                            cr = _call_claude_pdf(client, [chunk_path], cfg["prompt"], cfg["label"],
                                                  sleep_after=10, fallback=None)
                            if cr:
                                chunk_results.append(cr)
                    finally:
                        for chunk_path in chunk_paths:
                            chunk_path.unlink(missing_ok=True)
                    if chunk_results:
                        per_file.append(_merge_macro_research(chunk_results))
                else:
                    r = _call_claude_pdf(client, [pdf_path], cfg["prompt"], cfg["label"],
                                         sleep_after=10, fallback=None)
                    if r:
                        per_file.append(r)
            results["macro_research"]     = _merge_macro_research(per_file) if per_file else None
            cache_stats["macro_research"] = f"fresh ({len(recent_pdfs)} files)"

    return {
        "pdf":          results,
        "_pdf_sources": sources_used,
        "_cache_stats": cache_stats,
    }


# ── MAIN ──────────────────────────────────────────────────────────────────────

def _load_existing() -> dict | None:
    """Load macro_context.json from the last run, or None if missing/corrupt."""
    if not OUTPUT_PATH.exists():
        return None
    try:
        with open(OUTPUT_PATH, encoding='utf-8') as fh:
            return json.load(fh)
    except (json.JSONDecodeError, OSError):
        return None


def main():
    parser = argparse.ArgumentParser(description="Build macro_context.json")
    parser.add_argument(
        '--stage1-only', action='store_true',
        help='Skip all PDF extraction (no Claude API calls)',
    )
    parser.add_argument(
        '--force-pdf', action='store_true',
        help='Bypass PDF cache — re-extract every source even if unchanged',
    )
    args = parser.parse_args()

    if args.stage1_only:
        stage = "Stage 1 only (--stage1-only)"
    elif args.force_pdf:
        stage = "Stage 1 + 2 (--force-pdf, cache bypassed)"
    else:
        stage = "Stage 1 + 2"
    print(f"build_macro_context.py — {stage}")
    print(f"Output: {OUTPUT_PATH}\n")

    existing = None if args.stage1_only else _load_existing()

    etf = read_etf_pro()
    ham = read_ham_holdings()
    rta = read_rta_trades()
    rr  = read_risk_range_levels()

    now = datetime.now(tz=timezone.utc)

    sources_used = {
        "etf_pro":      etf.pop("_source"),
        "ham_holdings": ham.pop("_source"),
        "rta":          rta.pop("_source"),
        "risk_range":   rr.pop("_source"),
    }

    output = {
        "generated_at": now.isoformat(),
        "source_date":  date.today().isoformat(),
        "sources_used": sources_used,
        **etf,
        **ham,
        **rta,
        **rr,
    }

    cache_stats = {}
    if not args.stage1_only:
        pdf = extract_pdf_data(existing, args.force_pdf)
        cache_stats = pdf.pop("_cache_stats", {})
        sources_used.update(pdf.pop("_pdf_sources", {}))
        output.update(pdf)

    # sss_history — append today's SSS count on every full run, keep last 10
    sss_history = (existing or {}).get("sss_history", [])
    if not args.stage1_only:
        sss_data = output.get("pdf", {}).get("sss")
        if sss_data and sss_data.get("count") is not None:
            sss_history.append({
                "date":    date.today().isoformat(),
                "count":   sss_data["count"],
                "added":   len(sss_data.get("added") or []),
                "removed": len(sss_data.get("removed") or []),
            })
            sss_history = sss_history[-10:]
    output["sss_history"] = sss_history

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as fh:
        json.dump(output, fh, indent=2, ensure_ascii=False)

    rta_data = output.get('rta', {})
    pdf_data  = output.get('pdf', {})
    print(f"\n── Summary ──")
    print(f"  etf_rerank:              {len(output.get('etf_rerank', []))} tickers")
    print(f"  active_longs/shorts:     {len(output.get('active_longs', []))} / {len(output.get('active_shorts', []))}")
    print(f"  ham_holdings:            {len(output.get('ham_holdings', []))} tickers")
    print(f"  rta.recent_trades (90d): {len(rta_data.get('recent_trades', []))}")
    print(f"  rta.stats:               {rta_data.get('stats', {})}")
    print(f"  levels:                  {len(output.get('levels', {}))} tickers")
    if pdf_data:
        print()
        for k, v in pdf_data.items():
            stat   = cache_stats.get(k, '')
            status = f"OK ({stat})" if v is not None else f"null/failed ({stat})"
            print(f"  pdf.{k:<22} {status}")
    print(f"\nWritten: {OUTPUT_PATH}")


if __name__ == '__main__':
    main()
