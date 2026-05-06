#!/usr/bin/env python3
"""
import_official_levels.py
=========================
One-way data pipeline: reads the official Hedgeye risk range tracker Excel
workbook and writes a single JSON file for the website Risk Range tab.

What this script does:
  - Reads risk_range_tracker_excelworkbook.xlsx (read-only; never modified)
  - Scans all daily sheets to extract ticker / signal / buy / sell / close
  - Writes hedgeye-dashboard/data/official_levels.json
  - Exits early (no rewrite) if the workbook has not changed since last run

What this script does NOT do:
  - Does NOT call external APIs or download market data
  - Does NOT import or reference hedgeye_pvv_v2.py or Calibration_test.py
  - Does NOT use pandas
  - Does NOT modify the source workbook or any existing website files
  - Does NOT create any files other than official_levels.json
"""

import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed.")
    print("Run:  pip install openpyxl")
    sys.exit(1)

# ── PATHS ─────────────────────────────────────────────────────────────────────
SOURCE_PATH = Path(
    r"C:\Users\matth\OneDrive\Desktop\Trading"
    r"\risk_range_tracker_excelworkbook.xlsx"
)
OUTPUT_DIR = Path(
    r"C:\Users\matth\OneDrive\Desktop\Trading"
    r"\hedgeye-dashboard\data"
)
OUTPUT_PATH = OUTPUT_DIR / "official_levels.json"

# ── CONSTANTS ─────────────────────────────────────────────────────────────────
# Sheet names that are never daily data sheets
SKIP_SHEETS = frozenset({
    "RiskRanges", "Comparison", "Notes", "README",
    "Sheet1", "Sheet2", "Sheet3", "Summary", "Index", "Help",
})

VALID_SIGNALS = frozenset({"BULLISH", "BEARISH", "NEUTRAL"})

# Keyword sets for optional header-row column detection (lowercased)
_BUY_KW   = {"buy", "lrr", "low risk range", "buy trade", "support", "floor", "add"}
_SELL_KW  = {"sell", "trr", "top risk range", "sell trade", "resistance", "ceiling", "trim"}
_CLOSE_KW = {"close", "prev close", "previous close", "last", "price", "closing", "prev"}
_TICKER_KW = {"ticker", "symbol", "asset", "index"}
_SIGNAL_KW = {"signal", "trend", "direction", "bias"}
_AC_KW     = {"asset class", "class", "category", "type", "sector"}

# Ticker pattern: starts with ^ or uppercase letter, followed by
# uppercase letters / digits / slash / dot / dash / equals (up to 20 more chars)
# Examples: SPY, EUR/USD, USD/YEN, UST10Y, ^VIX, GBP/USD, BITCOIN
_TICKER_SIG_RE  = re.compile(r'^([\^A-Z][A-Z0-9/\.\-=]{0,20})\s*\(([A-Z]+)\)\s*$')
_TICKER_ONLY_RE = re.compile(r'^([\^A-Z][A-Z0-9/\.\-=]{0,20})$')


# ── UTILITIES ─────────────────────────────────────────────────────────────────

def get_mtime_iso(path: Path) -> str:
    """Return file last-modified time as UTC ISO-8601 string."""
    ts = os.path.getmtime(path)
    return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()


def sheet_name_to_date(name: str, year: int) -> Optional[str]:
    """Convert worksheet name to YYYY-MM-DD. Handles May 6. 2026 format."""
    import calendar
    s = name.strip()
    month_map = {mon.lower(): i for i, mon in enumerate(calendar.month_abbr) if mon}
    month_map.update({mon.lower(): i for i, mon in enumerate(calendar.month_name) if mon})
    pat = r"^([A-Za-z]+)[.]?\s+(\d{1,2})[.]?\s*(\d{4})$"
    mx = re.match(pat, s)
    if mx:
        mo = month_map.get(mx.group(1).lower())
        if mo:
            try:
                return datetime(int(mx.group(3)), mo, int(mx.group(2))).strftime("%Y-%m-%d")
            except ValueError:
                return None
    pat2 = r"(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})"
    mx = re.fullmatch(pat2, s)
    if mx:
        try:
            return datetime(int(mx.group(1)), int(mx.group(2)), int(mx.group(3))).strftime("%Y-%m-%d")
        except ValueError:
            return None
    pat3 = r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})"
    mx = re.fullmatch(pat3, s)
    if mx:
        mo2, dy, yr2 = int(mx.group(1)), int(mx.group(2)), int(mx.group(3))
        if yr2 < 100:
            yr2 += 2000
        try:
            return datetime(yr2, mo2, dy).strftime("%Y-%m-%d")
        except ValueError:
            return None
    return None


def cell_to_float(value) -> Optional[float]:
    """Convert a raw cell value to float, or None if non-numeric / missing."""
    if value is None:
        return None
    # Guard against Excel booleans slipping through isinstance(v, int)
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        f = float(value)
        return None if f != f else f        # reject NaN
    if isinstance(value, str):
        s = value.strip().replace(',', '').rstrip('%')
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def cell_to_str(value) -> str:
    """Convert a raw cell value to a stripped string."""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return ''   # date cells in wrong column — ignore
    return str(value).strip()


def smart_round(v: float) -> float:
    """
    Round to an appropriate number of decimal places based on magnitude.
      >= 100  : 2 dp  (equities, ETFs, gold prices)
      1–100   : 3 dp  (oil, FX rates like 27.xx)
      < 1     : 4 dp  (yields like 0.0425, FX like 0.xxx)
    """
    a = abs(v)
    if a >= 100:
        return round(v, 2)
    if a >= 1:
        return round(v, 3)
    return round(v, 4)


def parse_ticker_signal(raw: str):
    """
    Parse 'SPX (BULLISH)' -> ('SPX', 'BULLISH').
    Parse 'SPX'           -> ('SPX', None).
    Returns (None, None) if the string is not a ticker.
    """
    s = raw.strip()
    m = _TICKER_SIG_RE.match(s)
    if m:
        return m.group(1), m.group(2).upper()
    m = _TICKER_ONLY_RE.match(s)
    if m:
        return m.group(1), None
    return None, None


def detect_column_map(header_vals: list) -> dict:
    """
    Scan a potential header row and return {role: col_index}.
    Returns an empty dict if the row does not look like a header.
    Roles detected: ticker, buy, sell, close, signal, asset_class.
    """
    cols: dict = {}
    for i, v in enumerate(header_vals):
        h = cell_to_str(v).lower()
        if not h:
            continue
        if any(k in h for k in _TICKER_KW) and 'ticker' not in cols:
            cols['ticker'] = i
        elif any(k in h for k in _BUY_KW) and 'buy' not in cols:
            cols['buy'] = i
        elif any(k in h for k in _SELL_KW) and 'sell' not in cols:
            cols['sell'] = i
        elif any(k in h for k in _CLOSE_KW) and 'close' not in cols:
            cols['close'] = i
        elif any(k in h for k in _SIGNAL_KW) and 'signal' not in cols:
            cols['signal'] = i
        elif any(k in h for k in _AC_KW) and 'asset_class' not in cols:
            cols['asset_class'] = i
    return cols


# ── SHEET PARSER ──────────────────────────────────────────────────────────────

def parse_sheet(ws, sheet_date: str):
    """
    Parse one worksheet.

    Returns (rows, skipped_count) where each row is a dict:
      {ticker, date, close, buy, sell, signal, asset_class, name}

    skipped_count tracks rows with a recognisable ticker but missing / zero
    numeric values (buy / sell / close).

    Column layout strategy:
      1. Attempt header detection in the first 5 rows.
      2. Fall back to the positional layout used by the existing dashboard:
         Col 0 = "TICKER (SIGNAL)", Col 1 = buy/LRR, Col 2 = sell/TRR,
         Col 3 = prev-close.
    """
    parsed_rows: list = []
    skipped: int = 0

    # Materialise all non-blank rows as plain value lists
    all_rows: list = []
    for row in ws.iter_rows(values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        all_rows.append(list(row))

    if not all_rows:
        return parsed_rows, skipped

    # ── Column layout detection ───────────────────────────────────────────────
    col_map: dict = {}
    header_idx: Optional[int] = None

    for i, vals in enumerate(all_rows[:5]):
        detected = detect_column_map(vals)
        # Require at minimum: ticker + buy + sell
        if 'ticker' in detected and 'buy' in detected and 'sell' in detected:
            col_map = detected
            header_idx = i
            break

    # Fallback positional layout (matches existing dashboard column order)
    if not col_map:
        col_map = {'ticker': 0, 'buy': 1, 'sell': 2, 'close': 3}
        header_idx = None

    start   = (header_idx + 1) if header_idx is not None else 0
    t_col   = col_map.get('ticker', 0)
    b_col   = col_map.get('buy',    1)
    s_col   = col_map.get('sell',   2)
    c_col   = col_map.get('close',  3)
    sig_col = col_map.get('signal')         # may be absent
    ac_col  = col_map.get('asset_class')    # may be absent

    def _get(vals: list, idx: int):
        """Index into a row safely."""
        try:
            return vals[idx]
        except (IndexError, TypeError):
            return None

    # ── Data row iteration ────────────────────────────────────────────────────
    for i in range(start, len(all_rows)):
        vals = all_rows[i]

        ticker_raw = cell_to_str(_get(vals, t_col))
        if not ticker_raw:
            continue

        ticker, signal = parse_ticker_signal(ticker_raw)
        if ticker is None:
            # Not a ticker row (could be a name/description row) — skip
            continue

        # Signal from a dedicated column (used if not embedded in ticker text)
        if signal is None and sig_col is not None:
            sig_raw = cell_to_str(_get(vals, sig_col)).upper()
            if sig_raw in VALID_SIGNALS:
                signal = sig_raw

        if signal not in VALID_SIGNALS:
            skipped += 1
            continue

        buy_val   = cell_to_float(_get(vals, b_col))
        sell_val  = cell_to_float(_get(vals, s_col))
        close_val = cell_to_float(_get(vals, c_col))

        if buy_val is None or sell_val is None or close_val is None:
            skipped += 1
            continue
        if buy_val == 0 or sell_val == 0:
            skipped += 1
            continue

        asset_class = cell_to_str(_get(vals, ac_col)) if ac_col is not None else ''

        # Peek at the next row: if its first-column text does not match the
        # ticker pattern, treat it as a descriptive name for this ticker.
        name = ''
        if i + 1 < len(all_rows):
            nxt_raw = cell_to_str(_get(all_rows[i + 1], t_col))
            if nxt_raw and nxt_raw.lower() not in ('nan', ''):
                nt, _ = parse_ticker_signal(nxt_raw)
                if nt is None:
                    # Confirm it is not a bare number
                    try:
                        float(nxt_raw.replace(',', ''))
                    except ValueError:
                        name = nxt_raw

        parsed_rows.append({
            'ticker':      ticker,
            'date':        sheet_date,
            'close':       smart_round(close_val),
            'buy':         smart_round(buy_val),
            'sell':        smart_round(sell_val),
            'signal':      signal,
            'asset_class': asset_class,
            'name':        name,
        })

    return parsed_rows, skipped


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main() -> None:

    # 1. Verify source file exists
    if not SOURCE_PATH.exists():
        print("ERROR: Workbook not found at:")
        print(f"  {SOURCE_PATH}")
        print("Verify the path and try again.")
        sys.exit(1)

    # 2. Workbook modification timestamp (used for cache-busting)
    source_mtime = get_mtime_iso(SOURCE_PATH)
    year_hint    = datetime.fromisoformat(source_mtime).year

    print(f"Source  : {SOURCE_PATH.name}")
    print(f"Modified: {source_mtime}")

    # 3. Skip rewrite if workbook unchanged since last run
    if OUTPUT_PATH.exists():
        try:
            with open(OUTPUT_PATH, 'r', encoding='utf-8') as fh:
                existing = json.load(fh)
            if existing.get('source_modified') == source_mtime:
                print("No update needed — workbook unchanged.")
                return
        except (json.JSONDecodeError, KeyError, OSError):
            pass    # JSON missing or corrupt → fall through to regenerate

    # 4. Load workbook (data_only=True reads cached formula results, not formulas)
    print("Loading workbook…")
    try:
        wb = load_workbook(SOURCE_PATH, data_only=True)
    except Exception as exc:
        print(f"ERROR: Could not open workbook: {exc}")
        sys.exit(1)

    # 5. Iterate sheets
    sheets_scanned: list = []
    sheets_skipped: list = []
    all_rows:       list = []
    total_skipped:  int  = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip by name
        if sheet_name in SKIP_SHEETS:
            sheets_skipped.append(f"{sheet_name!r} — in skip list")
            continue

        # Skip hidden sheets
        try:
            if ws.sheet_state in ('hidden', 'veryHidden'):
                sheets_skipped.append(f"{sheet_name!r} — hidden")
                continue
        except AttributeError:
            pass    # attribute absent on some openpyxl versions; assume visible

        # Must parse as a date to be a daily data sheet
        sheet_date = sheet_name_to_date(sheet_name, year_hint)
        if sheet_date is None:
            sheets_skipped.append(f"{sheet_name!r} — not a date")
            continue

        rows, skip_count = parse_sheet(ws, sheet_date)
        total_skipped += skip_count

        if not rows:
            sheets_skipped.append(f"{sheet_name!r} — no valid rows")
            continue

        sheets_scanned.append(f"{sheet_name} -> {sheet_date}  ({len(rows)} rows)")
        all_rows.extend(rows)

    wb.close()

    if not all_rows:
        print("ERROR: No valid data rows found across all sheets.")
        sys.exit(1)

    # 6. Aggregate into per-ticker structure
    ticker_map: dict = {}

    for row in all_rows:
        t = row['ticker']
        if t not in ticker_map:
            ticker_map[t] = {
                'ticker':      t,
                'name':        row['name'],
                'asset_class': row['asset_class'],
                'history':     [],
            }
        else:
            # Backfill name / asset_class if found later
            if not ticker_map[t]['name'] and row['name']:
                ticker_map[t]['name'] = row['name']
            if not ticker_map[t]['asset_class'] and row['asset_class']:
                ticker_map[t]['asset_class'] = row['asset_class']

        ticker_map[t]['history'].append({
            'date':   row['date'],
            'close':  row['close'],
            'buy':    row['buy'],
            'sell':   row['sell'],
            'signal': row['signal'],
        })

    # Sort history chronologically; deduplicate (last entry wins per date)
    for t, data in ticker_map.items():
        deduped: dict = {}
        for h in data['history']:
            deduped[h['date']] = h      # later occurrence overwrites earlier
        data['history'] = sorted(deduped.values(), key=lambda x: x['date'])
        data['latest']  = data['history'][-1]

    # 7. Top-level metadata
    all_dates = sorted({
        h['date']
        for data in ticker_map.values()
        for h in data['history']
    })
    latest_date       = all_dates[-1] if all_dates else ''
    latest_date_count = sum(
        1 for data in ticker_map.values()
        if data['latest']['date'] == latest_date
    )
    total_rows = sum(len(data['history']) for data in ticker_map.values())

    # 8. Write JSON
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    output = {
        'generated_at':    datetime.now(tz=timezone.utc).isoformat(),
        'source_file':     SOURCE_PATH.name,
        'source_modified': source_mtime,
        'ticker_count':    len(ticker_map),
        'row_count':       total_rows,
        'latest_date':     latest_date,
        'tickers':         ticker_map,
    }

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as fh:
        json.dump(output, fh, indent=2, ensure_ascii=False)

    # 9. Success summary
    print(
        f"\nWritten: {len(ticker_map)} tickers, {total_rows} rows, "
        f"latest date {latest_date} -> data/official_levels.json"
    )

    # 10. Validation summary
    print("\nValidation summary")
    print(f"  Sheets scanned           : {len(sheets_scanned)}")
    print(f"  Sheets skipped           : {len(sheets_skipped)}")
    print(f"  Tickers found            : {len(ticker_map)}")
    print(f"  Total valid rows         : {total_rows}")
    print(f"  Latest date              : {latest_date}")
    print(f"  Latest date ticker count : {latest_date_count}")
    print(f"  Rows skipped (bad data)  : {total_skipped}")

    if sheets_scanned:
        print(f"\n  Sheets scanned ({len(sheets_scanned)}):")
        for s in sheets_scanned:
            print(f"    {s}")

    if sheets_skipped:
        print(f"\n  Sheets skipped ({len(sheets_skipped)}):")
        for s in sheets_skipped:
            print(f"    {s}")


if __name__ == '__main__':
    main()
